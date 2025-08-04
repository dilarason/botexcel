# converter.py
# Ham veri çıkarım ve hata/log yönetimi eklenmiş hali

import mimetypes           # Dosya türünü belirlemek için
import fitz                # PyMuPDF, PDF okumak için
from PIL import Image      # Görsel işleme için
import pytesseract         # OCR (görselden metin) için
from barcode_module import handle_barcode   # Barkod okuma modülü
import logging             # Loglama için
import os                  # Klasör kontrolü için
import csv                 # CSV için
from docx import Document  # DOCX için
import json                # JSON için 
from openpyxl import load_workbook # XLSX için  

# Log klasörü varsa yoksa oluştur
os.makedirs('logs', exist_ok=True)

# Loglama ayarları: tüm hatalar logs/app.log'a gider
logging.basicConfig(
    filename='logs/app.log',
    filemode='a',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def convert_input(file_path):
    """
    Her türlü girdi dosyasını (PDF, TXT, Görsel) okur, hataları loglar.
    Hatalı durumlarda boş dizi döner.
    """
    try:
        # Dosya türünü belirle
        mime_type, _ = mimetypes.guess_type(file_path)
        data = []

        if mime_type == "application/pdf":
            try:
                # PDF'deki barkodları bul
                barcodes = handle_barcode(file_path)
                data.extend(f"BARCODE: {c}" for c in barcodes)

                # PDF sayfa sayfa oku
                doc = fitz.open(file_path)
                for page in doc:
                    text = page.get_text().strip()
                    if text:
                        data.append(text)
                doc.close()
                logging.info(f"{file_path} PDF olarak başarıyla işlendi.")
            except Exception as e:
                logging.error(f"{file_path} PDF işlenirken hata: {e}", exc_info=True)

        elif mime_type == "text/plain":
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if line:
                            data.append(line)
                logging.info(f"{file_path} TXT olarak başarıyla işlendi.")
            except Exception as e:
                logging.error(f"{file_path} TXT işlenirken hata: {e}", exc_info=True)

        elif mime_type and mime_type.startswith("image/"):
            try:
                # Görseldeki barkodları bul
                barcodes = handle_barcode(file_path)
                data.extend(f"BARCODE: {c}" for c in barcodes)

                # Görselden metin çıkar (OCR)
                img = Image.open(file_path)
                text = pytesseract.image_to_string(img, lang="tur")
                for line in text.splitlines():
                    line = line.strip()
                    if line:
                        data.append(line)
                logging.info(f"{file_path} görsel olarak başarıyla işlendi.")
            except Exception as e:
                logging.error(f"{file_path} görsel işlenirken hata: {e}", exc_info=True)

        elif mime_type == "text/csv":
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        # Satırı stringe çevirip ekle
                        line = ", ".join(row).strip()
                        if line:
                            data.append(line)
                logging.info(f"{file_path} CSV olarak başarıyla işlendi.")
            except Exception as e:
                logging.error(f"{file_path} CSV işlenirken hata: {e}", exc_info=True)

        elif mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            try:
                doc = Document(file_path)
                for para in doc.paragraphs:
                    text = para.text.strip()
                    if text:
                        data.append(text)
                logging.info(f"{file_path} DOCX olarak başarıyla işlendi.")
            except Exception as e:
                logging.error(f"{file_path} DOCX işlenirken hata: {e}", exc_info=True)

        elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            try:
                wb = load_workbook(file_path, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    line = ", ".join([str(cell) if cell is not None else "" for cell in row]).strip()
                    if line:
                        data.append(line)
                wb.close()
                logging.info(f"{file_path} XLSX olarak başarıyla işlendi.")
            except Exception as e:
                logging.error(f"{file_path} XLSX işlenirken hata: {e}", exc_info=True)

        elif mime_type == "application/json":
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    obj = json.load(f)
                    # Satırlaşmış gösterim
                    def flatten(o, prefix=""):
                        if isinstance(o, dict):
                            for k, v in o.items():
                                yield from flatten(v, f"{prefix}{k}.")
                        elif isinstance(o, list):
                            for idx, v in enumerate(o):
                                yield from flatten(v, f"{prefix}{idx}.")
                        else:
                            yield f"{prefix[:-1]}: {o}"
                    for line in flatten(obj):
                        data.append(line)
                logging.info(f"{file_path} JSON olarak başarıyla işlendi.")
            except Exception as e:
                logging.error(f"{file_path} JSON işlenirken hata: {e}", exc_info=True)

        else:
            logging.warning(f"{file_path} desteklenmeyen dosya türü: {mime_type}")

        return data

    except Exception as e:
        # Beklenmedik her hata buraya düşer
        logging.error(f"convert_input beklenmeyen hata: {e}", exc_info=True)
        return []
